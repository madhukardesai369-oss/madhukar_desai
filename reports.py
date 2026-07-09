
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import mm
from storage import load_bookings
from analytics import summary_stats

OUT_DIR = Path(__file__).parent / "outputs"
OUT_DIR.mkdir(exist_ok=True)


def export_excel():
    bookings = load_bookings()
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    stats = summary_stats()
    ws["A1"] = "Hotel Booking Report"
    ws["A1"].font = Font(size=16, bold=True, color="1a365d")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    row = 4
    for k, v in stats.items():
        ws.cell(row=row, column=1, value=k.replace("_", " ").title()).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1

    # Sheet 2: Bookings
    ws2 = wb.create_sheet("Bookings")
    headers = ["Booking ID", "Customer", "Age", "Mobile", "ID Proof", "Address",
               "Room", "Category", "Check-in", "Check-out", "Nights",
               "Total (INR)", "Status", "Created At", "Created By"]
    ws2.append(headers)
    header_fill = PatternFill("solid", fgColor="2b6cb0")
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for b in bookings:
        c = b["customer"]
        ws2.append([b["booking_id"], c["name"], c["age"], c["mobile"], c["id_proof"],
                    c["address"], b["room_number"], b["category"],
                    b["check_in"], b["check_out"], b["nights"],
                    b["total_amount"], b["status"], b["created_at"],
                    b.get("created_by", "")])
    for col_cells in ws2.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)

    path = OUT_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return str(path)


def export_pdf():
    bookings = load_bookings()
    path = OUT_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "t",
        parent=styles["Title"],
        textColor=colors.HexColor("#1a365d"),
        fontSize=20,
        alignment=1,
    )
    story = [
        Paragraph("Hotel Booking Report", title),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                  styles["Italic"]),
        Spacer(1, 6 * mm)
    ]

    stats = summary_stats()
    stats_tbl = Table(
        [[k.replace("_", " ").title(), str(v)] for k, v in stats.items()],
        colWidths=[60 * mm, 40 * mm]
    )
    stats_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#edf2f7")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(stats_tbl)
    story.append(Spacer(1, 6 * mm))

    rows = [["ID", "Customer", "Mobile", "Room", "Cat", "In", "Out", "N",
             "Total", "Status"]]
    for b in bookings:
        rows.append([b["booking_id"], b["customer"]["name"], b["customer"]["mobile"],
                     b["room_number"], b["category"], b["check_in"], b["check_out"],
                     b["nights"], f"{b['total_amount']:.0f}", b["status"]])
    tbl = Table(rows, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b6cb0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.whitesmoke, colors.white]),
    ]))
    story.append(tbl)
    doc.build(story)
    return str(path)